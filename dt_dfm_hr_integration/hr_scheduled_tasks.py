from ftplib import FTP
import frappe
from frappe.model.document import Document
import openpyxl
from io import BytesIO
from openpyxl import load_workbook, workbook
from openpyxl.writer.excel import save_virtual_workbook



import paramiko
from io import BytesIO
import openpyxl
import frappe
import pysftp



def all():
    pass






# def cron():
#     date = frappe.utils.now_datetime()
#     today_date = date.date().day

#     if (today_date == 1) or (today_date == 2) or (today_date == 3):
#         print("\n\n\nDFM HR Integration Process Start\n\n\n")

#         settings = frappe.get_single("DFM HR Settings")

#         server_address = settings.sftp_server_address
#         user = settings.sftp_user
#         password = settings.get_password('sftp_password')
#         port = settings.sftp_port

#         folder = settings.sftp_folder

#         ftp = FTP()
#         ftp.connect(server_address, port)
#         ftp.login(user=user, passwd=password)
#         ftp.set_pasv(False)

#         if folder:
#             ftp.cwd(folder)
#             file_list = ftp.nlst()
#         else:
#             file_list = ftp.nlst()


#         xlsx_files = [file for file in file_list if file.lower().endswith(".xlsx")]

#         existing_file_names = frappe.get_all("DFM HR Log", filters={}, fields=["file_name"])
#         existing_file_names = [file["file_name"] for file in existing_file_names]

#         for file_name in xlsx_files:
#             if file_name in existing_file_names:
#                 print("File name {} already exists in DFM HR Log. Skipping file...".format(file_name))
#                 continue

#             try:
#                 file_content = []
#                 ftp.retrbinary('RETR ' + file_name, file_content.append)
#                 print("Processing file: {}".format(file_name))


#                 file_content_io = BytesIO(b''.join(file_content))

#                 workbook = openpyxl.load_workbook(file_content_io)
#                 sheet = workbook.active

                
                
                
#                 file_doc = None

#                 if not frappe.get_all("File", filters={"file_name": file_name}):
#                     file_content_str = save_virtual_workbook(workbook)
#                     file_doc = frappe.get_doc({
#                         "doctype": "File",
#                         "file_name": file_name,
#                         "content": file_content_str,
#                         "is_private": 1,
#                         "folder": "Home"
#                     })
#                     file_doc.insert(ignore_permissions=True)
#                 else:
#                     file_doc = frappe.get_doc("File", {"file_name": file_name})




#                 header_row = [cell.value for cell in sheet[2]]
#                 print("Header Row: {}".format(header_row))

#                 # for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
#                 #     if not all(cell is None for cell in row):
#                 #         try:
#                 #             create_salary_register_entry(row, header_row, file_name, row_number, file_doc)
#                 #         except Exception as e:
#                 #             log_error(file_name, row_number, file_doc, str(e))


#                 grouped_rows = {}
#                 for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
#                     if not all(cell is None for cell in row):
#                         companies_value = row[header_row.index('Companies')]
#                         if companies_value not in grouped_rows:
#                             grouped_rows[companies_value] = []
#                         grouped_rows[companies_value].append((row_number, row))

#                 # Print the grouped rows
#                 for companies_value, rows in grouped_rows.items():
#                     create_salary_register_entry(rows, header_row, companies_value, file_name, file_doc)
#                     # print(f"Rows with Companies value '{companies_value}':")
#                     # print(rows)

#                     # for row_number, row in rows:
#                     #     print(f"  Row {row_number}: {row}")

#             except Exception as e:
#                 log_error(file_name, "", "", str(e))
#                 continue
#         ftp.quit()
#         print("\n\n\n\n\nAll files have been downloaded and processed.")
#         frappe.msgprint("Successfully synced data from SFTP.")
    
#     else:
#         pass














# def cron():
#     date = frappe.utils.now_datetime()
#     today_date = date.date().day

#     if 1 <= today_date <= 3:
#         try:
#             print("DFM HR Integration Process Start")

#             settings = frappe.get_single("DFM HR Settings")

#             server_address = settings.sftp_server_address
#             user = settings.sftp_user
#             password = settings.get_password('sftp_password')
#             port = settings.sftp_port

#             folder = settings.sftp_folder

#             cnopts = pysftp.CnOpts()
#             cnopts.hostkeys = None  # Disable host key checking (use with caution)

#             with pysftp.Connection(server_address, port=port, username=user, password=password, cnopts=cnopts) as sftp:
#                 if folder:
#                     sftp.chdir(folder)
#                     file_list = sftp.listdir()
#                 else:
#                     file_list = sftp.listdir()

#                 xlsx_files = [file for file in file_list if file.lower().endswith(".xlsx")]

#                 existing_file_names = frappe.get_all("DFM HR Log", filters={}, fields=["file_name"])
#                 existing_file_names = [file["file_name"] for file in existing_file_names]

#                 for file_name in xlsx_files:
#                     if file_name in existing_file_names:
#                         print("File name {} already exists in DFM HR Log. Skipping file...".format(file_name))
#                         continue

#                     try:
#                         # Retrieve the content as bytes
#                         with sftp.open(file_name) as f:
#                             file_content = f.read()

#                         print("Processing file: {}".format(file_name))

#                         # Use BytesIO for the content
#                         file_content_io = BytesIO(file_content)

#                         workbook = openpyxl.load_workbook(file_content_io)
#                         sheet = workbook.active

#                         file_doc = None

#                         if not frappe.get_all("File", filters={"file_name": file_name}):
#                             file_content_str = save_virtual_workbook(workbook)
#                             file_doc = frappe.get_doc({
#                                 "doctype": "File",
#                                 "file_name": file_name,
#                                 "content": file_content_str,
#                                 "is_private": 1,
#                                 "folder": "Home"
#                             })
#                             file_doc.insert(ignore_permissions=True)
#                         else:
#                             file_doc = frappe.get_doc("File", {"file_name": file_name})

#                         header_row = [cell.value for cell in sheet[2]]
#                         print("Header Row: {}".format(header_row))

#                         grouped_rows = {}
#                         for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
#                             if not all(cell is None for cell in row):
#                                 companies_value = row[header_row.index('Business Unit')]
#                                 company = frappe.get_value("DFM HR Company", 
#                                                 filters= {
#                                                     "business_unit" : companies_value
#                                                 },
#                                                 fieldname = "company")
                                
#                                 if company not in grouped_rows:
#                                     grouped_rows[company] = []
#                                 grouped_rows[company].append((row_number, row))

#                         for company, rows in grouped_rows.items():
#                             create_salary_register_entry(rows, header_row, company, file_name, file_doc)

#                     except Exception as e:
#                         log_error(file_name, "", "", str(e))
#                         continue

#             print("All files have been downloaded and processed.")
#             # frappe.msgprint("Successfully synced data from SFTP.")

#         except Exception as e:
#             log_error("", "", "", str(e))
#             # frappe.msgprint("Error occurred during SFTP sync. Check logs for details.")

#     else:
#         pass


















def cron():
    print("\n\n\nDFM HR Integration Process Start\n\n\n")

    settings = frappe.get_single("DFM HR Settings")

    server_address = settings.sftp_server_address
    user = settings.sftp_user
    password = settings.get_password('sftp_password')
    port = settings.sftp_port

    folder = settings.sftp_folder

    cnopts = pysftp.CnOpts()
    cnopts.hostkeys = None  # Disable host key checking (use with caution)

    with pysftp.Connection(server_address, port=port, username=user, password=password, cnopts=cnopts) as sftp:
        if folder:
            sftp.chdir(folder)
            file_list = sftp.listdir()
        else:
            file_list = sftp.listdir()

        xlsx_files = [file for file in file_list if file.lower().endswith(".xlsx")]

        existing_file_names = frappe.get_all("DFM HR Log", filters={}, fields=["file_name"])
        existing_file_names = [file["file_name"] for file in existing_file_names]

        for file_name in xlsx_files:
            if file_name in existing_file_names:
                print("File name {} already exists in DFM HR Log. Skipping file...".format(file_name))
                continue

            try:
                # Retrieve the content as bytes
                with sftp.open(file_name) as f:
                    file_content = f.read()

                print("Processing file: {}".format(file_name))

                # Use BytesIO for the content
                file_content_io = BytesIO(file_content)

                workbook = openpyxl.load_workbook(file_content_io)
                sheet = workbook.active

                file_doc = None

                if not frappe.get_all("File", filters={"file_name": file_name}):
                    file_content_str = save_virtual_workbook(workbook)
                    file_doc = frappe.get_doc({
                        "doctype": "File",
                        "file_name": file_name,
                        "content": file_content_str,
                        "is_private": 1,
                        "folder": "Home"
                    })
                    file_doc.insert(ignore_permissions=True)
                else:
                    file_doc = frappe.get_doc("File", {"file_name": file_name})

                header_row = [cell.value for cell in sheet[2]]
                print("Header Row: {}".format(header_row))

                grouped_rows = {}
                for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                    if not all(cell is None for cell in row):
                        companies_value = row[header_row.index('Business Unit')]
                        company = frappe.get_value("DFM HR Company", 
                                          filters= {
                                              "business_unit" : companies_value
                                          },
                                          fieldname = "company")
                        
                        if company not in grouped_rows:
                            grouped_rows[company] = []
                        grouped_rows[company].append((row_number, row))

                for company, rows in grouped_rows.items():
                    create_salary_register_entry(rows, header_row, company, file_name, file_doc)

            except Exception as e:
                log_error(file_name, "", "", str(e))
                continue

    print("\n\n\n\n\nAll files have been downloaded and processed.")
    # frappe.msgprint("Successfully synced data from SFTP.")




















def create_salary_register_entry(rows, header_row, company, file_name, file_doc):
    try:
        salary_register = frappe.new_doc('DFM HR Salary Transaction Summary')

        # company = frappe.get_value("DFM HR Company", 
        #                                   filters= {
        #                                       "business_unit" : companies_value
        #                                   },
        #                                   fieldname = "company")
        
        salary_register.company = company
        salary_register.date = frappe.utils.now_datetime()

        details_list = []

        debit_amount = 0
        credit_amount = 0

        for row_number, row in rows:

            department = row[header_row.index('Department')]
            dfm_hr_grade = row[header_row.index('Grade')]
            

            for header in header_row:
                amount = row[header_row.index(header)]

                # Skip appending if the amount is 0 or not a number
                if isinstance(amount, (int, float)) and amount != 0:
                    if frappe.get_all("DFM HR Salary Head", filters={"salary_head": header}):
                        gl_mapping = frappe.get_all("DFM HR GL Mapping",
                                                    filters={
                                                        "company": company,
                                                        "department": department,
                                                        "dfm_hr_grade": dfm_hr_grade
                                                    })

                        if gl_mapping:
                            account = frappe.get_value("DFM HR GL Details",
                                                    filters={
                                                        "parent": gl_mapping[0].name,
                                                        "salary_head": header
                                                    },
                                                    fieldname="account")
                            
                            type = frappe.get_value("DFM HR GL Details",
                                                    filters={
                                                        "parent": gl_mapping[0].name,
                                                        "salary_head": header
                                                    },
                                                    fieldname="type")
                        
                            if account and type:
                                details_list.append({
                                    "company": company,
                                    "dfm_hr_grade": dfm_hr_grade,
                                    "department": department,
                                    "salary_head": header,
                                    "account": account,
                                    "amount": amount,
                                    "type": type 
                                })

                                # Update debit_amount or credit_amount based on the type
                                if type == "Debit":
                                    debit_amount += amount
                                elif type == "Credit":
                                    credit_amount += amount

        salary_register.set("dfm_hr_salary_transaction_summary_details", details_list)

        salary_register.total_debit = debit_amount
        salary_register.total_credit = credit_amount

        salary_register.insert(ignore_permissions=True)
        salary_register.submit()

        # Create Journal Entry
        create_journal_entry(salary_register, details_list, file_name, rows, file_doc)

    except Exception as e:
        log_error(file_name, rows, file_doc, str(e))















def create_journal_entry(salary_register, details_list, file_name, rows, file_doc):
    try:
        journal_entry = frappe.new_doc("Journal Entry")
        journal_entry.voucher_type = "Journal Entry"
        journal_entry.posting_date = frappe.utils.nowdate()
        journal_entry.company = salary_register.company
        journal_entry.user_remark = f"Salary Entry"
        journal_entry.dfm_hr_salary_transaction_summary = salary_register.name

        cost_center = frappe.db.get_value('Company', salary_register.company, 'cost_center')

        account_totals = {}  # Dictionary to store accumulated amounts for each account

        for entry in details_list:
            amount = entry.get("amount")
            account = entry.get("account")
            entry_type = entry.get("type")

            # Initialize the account entry in the dictionary if not present
            if account not in account_totals:
                account_totals[account] = {"Debit": 0, "Credit": 0}

            # Accumulate amounts based on the type (Debit or Credit)
            if entry_type == "Debit":
                account_totals[account]["Debit"] += amount
            elif entry_type == "Credit":
                account_totals[account]["Credit"] += amount

        total_debit = 0
        total_credit = 0

        # Iterate through the accumulated amounts and add entries to the journal entry
        for account, amounts in account_totals.items():
            debit_amount = amounts["Debit"]
            credit_amount = amounts["Credit"]

            if debit_amount > 0:
                journal_entry.append("accounts", {
                    "account": account,
                    "cost_center": cost_center,
                    "debit_in_account_currency": debit_amount,
                })
                total_debit += debit_amount

            if credit_amount > 0:
                journal_entry.append("accounts", {
                    "account": account,
                    "cost_center": cost_center,
                    "credit_in_account_currency": credit_amount,
                })
                total_credit += credit_amount

        # Ensure total_credit matches total_debit
        if total_credit != total_debit:
            frappe.throw("Total Credit should match Total Debit in the Journal Entry.")

        journal_entry.total_debit = total_debit
        journal_entry.total_credit = total_credit
        journal_entry.insert(ignore_permissions=True)
        journal_entry.submit()

        log_success(file_name, rows, salary_register.name, journal_entry.name, file_doc)

    except Exception as e:
        log_partial_success(file_name, rows, salary_register, file_doc, str(e))









# def create_journal_entry(salary_register, details_list, file_name, rows, file_doc):
#     try:
#         journal_entry = frappe.new_doc("Journal Entry")
#         journal_entry.voucher_type = "Journal Entry"
#         journal_entry.posting_date = frappe.utils.nowdate()
#         journal_entry.company = salary_register.company
#         journal_entry.user_remark = f"Salary Entry"
#         journal_entry.dfm_hr_salary_transaction_summary = salary_register.name

#         cost_center = frappe.db.get_value('Company', salary_register.company, 'cost_center')

#         total_debit = 0
#         total_credit = 0

#         for entry in details_list:
#             amount = entry.get("amount")
#             account = entry.get("account")
#             entry_type = entry.get("type")

#             if entry_type == "Debit":
#                 journal_entry.append("accounts", {
#                     "account": account,
#                     "cost_center": cost_center,
#                     "debit_in_account_currency": amount,
#                 })
#                 total_debit += amount

#             elif entry_type == "Credit":
#                 journal_entry.append("accounts", {
#                     "account": account,
#                     "cost_center": cost_center,
#                     "credit_in_account_currency": amount,
#                 })
#                 total_credit += amount

#         # Ensure total_credit matches total_debit
#         if total_credit != total_debit:
#             frappe.throw("Total Credit should match Total Debit in the Journal Entry.")

#         journal_entry.total_debit = total_debit
#         journal_entry.total_credit = total_credit
#         journal_entry.insert(ignore_permissions=True)
#         journal_entry.submit()

#         log_success(file_name, rows, salary_register.name, journal_entry.name, file_doc)

#     except Exception as e:
#         log_partial_success(file_name, rows, salary_register, file_doc, str(e))







def log_success(file_name, rows, salary_register, journal_entry, file_doc):
    log_entry = frappe.new_doc('DFM HR Log')
    log_entry.file_name = file_name
    log_entry.file_path = file_doc.file_url

    row_numbers = [row_number for row_number, _ in rows]
    log_entry.row_number = ', '.join(map(str, row_numbers))

    log_entry.status = "Success"
    log_entry.dfm_hr_salary_transaction_summary = salary_register
    log_entry.journal_entry = journal_entry
    log_entry.date = frappe.utils.now_datetime()
    log_entry.error = ""
    log_entry.insert(ignore_permissions=True)


def log_partial_success(file_name, rows, salary_register, file_doc, error):
    log_entry = frappe.new_doc('DFM HR Log')
    log_entry.file_name = file_name
    log_entry.file_path = file_doc.file_url

    row_numbers = [row_number for row_number, _ in rows]
    log_entry.row_number = ', '.join(map(str, row_numbers))

    log_entry.status = "Partial Success"
    log_entry.dfm_hr_salary_transaction_summary = salary_register
    log_entry.date = frappe.utils.now_datetime()
    log_entry.error = error
    log_entry.insert(ignore_permissions=True)



def log_error(file_name, rows, file_doc, error):
    log_entry = frappe.new_doc('DFM HR Log')
    log_entry.file_name = file_name

    if file_doc:
        log_entry.file_path = file_doc.file_url
    
    if rows:
        row_numbers = [row_number for row_number, _ in rows]
        log_entry.row_number = ', '.join(map(str, row_numbers))

    log_entry.status = "Error"
    log_entry.date = frappe.utils.now_datetime()
    log_entry.error = error
    log_entry.insert(ignore_permissions=True)





