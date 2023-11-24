from ftplib import FTP
import frappe
from frappe.model.document import Document
import openpyxl
from io import BytesIO
from openpyxl import workbook
from openpyxl.writer.excel import save_virtual_workbook



import paramiko
from io import BytesIO
import openpyxl
import frappe



def all():
    pass



# def cron():
#     date = frappe.utils.now_datetime()
#     today_date = date.date().day

#     if (today_date == 1) or (today_date == 2) or (today_date == 3):

#         print("\n\n\nDFM HR Integration Process Start\n\n\n")

#         settings = frappe.get_single("DFM HR Settings")

#         server_address = settings.sftp_server_address
#         port = settings.sftp_port
#         user = settings.sftp_user
#         password = settings.get_password('sftp_password')

#         # Establish an SFTP connection
#         transport = paramiko.Transport((server_address, port))
#         transport.connect(username=user, password=password)
#         sftp = paramiko.SFTPClient.from_transport(transport)

#         file_list = sftp.listdir()

#         xlsx_files = [file for file in file_list if file.lower().endswith(".xlsx")]

#         for file_name in xlsx_files:
#             try:
#                 file_content = []
#                 with sftp.file(file_name, 'rb') as remote_file:
#                     file_content_io = BytesIO(remote_file.read())
#                     workbook = openpyxl.load_workbook(file_content_io)
#                 sheet = workbook.active

                
                
                
#                 file_doc = None

#                 if not frappe.get_all("File", filters={"file_name": file_name}):
#                     file_content_str = save_virtual_workbook(workbook)
#                     file_doc = frappe.get_doc({
#                         "doctype": "File",
#                         "file_name": file_name,
#                         "content": file_content_str,
#                         "is_private": 1,
#                         "folder": "Home"
#                     })
#                     file_doc.insert(ignore_permissions=True)
#                 else:
#                     file_doc = frappe.get_doc("File", {"file_name": file_name})




#                 header_row = [cell.value for cell in sheet[2]]
#                 print("Header Row: {}".format(header_row))

#                 for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
#                     if not all(cell is None for cell in row):
#                         try:
#                             create_salary_register_entry(row, header_row, file_name, row_number, file_doc)
#                         except Exception as e:
#                             log_error(file_name, row_number, file_doc, str(e))

#             except Exception as e:
#                 log_error(file_name, "", str(e))
#                 continue

#         sftp.close()
#         transport.close()

#         print("\n\n\n\n\nAll files have been downloaded and processed.")
#         frappe.msgprint("Successfully synced data from SFTP.")
    
#     else:
#         pass
    















def cron():
    date = frappe.utils.now_datetime()
    today_date = date.date().day

    if (today_date == 1) or (today_date == 2) or (today_date == 3):
        print("\n\n\nDFM HR Integration Process Start\n\n\n")

        settings = frappe.get_single("DFM HR Settings")

        server_address = settings.sftp_server_address
        user = settings.sftp_user
        password = settings.get_password('sftp_password')


        ftp = FTP(server_address)
        ftp.login(user=user, passwd=password)
        ftp.set_pasv(False)

        file_list = ftp.nlst()

        xlsx_files = [file for file in file_list if file.lower().endswith(".xlsx")]

        for file_name in xlsx_files:
            try:
                file_content = []
                ftp.retrbinary('RETR ' + file_name, file_content.append)
                print("Processing file: {}".format(file_name))


                file_content_io = BytesIO(b''.join(file_content))

                workbook = openpyxl.load_workbook(file_content_io)
                sheet = workbook.active

                
                
                
                file_doc = None

                if not frappe.get_all("File", filters={"file_name": file_name}):
                    file_content_str = save_virtual_workbook(workbook)
                    file_doc = frappe.get_doc({
                        "doctype": "File",
                        "file_name": file_name,
                        "content": file_content_str,
                        "is_private": 1,
                        "folder": "Home"
                    })
                    file_doc.insert(ignore_permissions=True)
                else:
                    file_doc = frappe.get_doc("File", {"file_name": file_name})




                header_row = [cell.value for cell in sheet[2]]
                print("Header Row: {}".format(header_row))

                for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                    if not all(cell is None for cell in row):
                        try:
                            create_salary_register_entry(row, header_row, file_name, row_number, file_doc)
                        except Exception as e:
                            log_error(file_name, row_number, file_doc, str(e))

            except Exception as e:
                log_error(file_name, "", str(e))
                continue
        ftp.quit()
        print("\n\n\n\n\nAll files have been downloaded and processed.")
        frappe.msgprint("Successfully synced data from SFTP.")
    
    else:
        pass









def create_salary_register_entry(row, header_row, file_name, row_number, file_doc):
    try:
        salary_register = frappe.new_doc('DFM HR Salary Transaction Summary')
        salary_register.department = row[header_row.index('Department')]
        salary_register.dfm_hr_grade = row[header_row.index('Grade')]
        salary_register.company = row[header_row.index('Companies')]
        salary_register.date = frappe.utils.now_datetime()


        details_list = []

        for header in header_row:
            amount = row[header_row.index(header)]

            # Skip appending if the amount is 0 or not a number
            if isinstance(amount, (int, float)) and amount != 0:
                
                if frappe.get_all("DFM HR Salary Head", filters={"salary_head": header}):
                    
                    gl_mapping = frappe.get_all("DFM HR GL Mapping",
                                                filters={
                                                    "company": salary_register.company,
                                                    "department": salary_register.department,
                                                    "dfm_hr_grade": salary_register.dfm_hr_grade
                                                })

                    if gl_mapping:
                        account = frappe.get_value("DFM HR GL Details",
                                                filters={
                                                    "parent": gl_mapping[0].name,
                                                    "salary_head": header
                                                },
                                                fieldname="account")
                        
                        type = frappe.get_value("DFM HR GL Details",
                                                filters={
                                                    "parent": gl_mapping[0].name,
                                                    "salary_head": header
                                                },
                                                fieldname="type")
                    

                        if account and type:
                            details_list.append({
                                "salary_head": header,
                                "account": account,
                                "amount": amount,
                                "type": type 
                            })

        salary_register.set("dfm_hr_salary_transaction_summary_details", details_list)

        total_amount = sum(entry["amount"] for entry in details_list)
        salary_register.total_amount = total_amount

        salary_register.insert(ignore_permissions=True)
        salary_register.submit()

        # Create Journal Entry
        create_journal_entry(salary_register, details_list, file_name, row_number, file_doc)

    except Exception as e:
        log_error(file_name, row_number, file_doc, str(e))










def create_journal_entry(salary_register, details_list, file_name, row_number, file_doc):
    try:
        journal_entry = frappe.new_doc("Journal Entry")
        journal_entry.voucher_type = "Journal Entry"
        journal_entry.posting_date = frappe.utils.nowdate()
        journal_entry.company = salary_register.company
        journal_entry.user_remark = f"Salary Entry"
        journal_entry.dfm_hr_salary_transaction_summary = salary_register.name

        total_debit = 0
        total_credit = 0

        for entry in details_list:
            amount = entry.get("amount")
            account = entry.get("account")
            entry_type = entry.get("type")

            if entry_type == "Debit":
                journal_entry.append("accounts", {
                    "account": account,
                    "debit_in_account_currency": amount,
                })
                total_debit += amount

            elif entry_type == "Credit":
                journal_entry.append("accounts", {
                    "account": account,
                    "credit_in_account_currency": amount,
                })
                total_credit += amount

        # Ensure total_credit matches total_debit
        if total_credit != total_debit:
            frappe.throw("Total Credit should match Total Debit in the Journal Entry.")

        journal_entry.total_debit = total_debit
        journal_entry.total_credit = total_credit
        journal_entry.insert(ignore_permissions=True)
        journal_entry.submit()

        log_success(file_name, row_number, salary_register.name, journal_entry.name, file_doc)

    except Exception as e:
        log_partial_success(file_name, row_number, salary_register, file_doc, str(e))



def log_success(file_name, row_number, salary_register, journal_entry, file_doc):
    log_entry = frappe.new_doc('DFM HR Log')
    log_entry.file_name = file_name
    log_entry.file_path = file_doc.file_url
    log_entry.row_number = row_number
    log_entry.status = "Success"
    log_entry.dfm_hr_salary_transaction_summary = salary_register
    log_entry.journal_entry = journal_entry
    log_entry.date = frappe.utils.now_datetime()
    log_entry.error = ""
    log_entry.insert(ignore_permissions=True)


def log_partial_success(file_name, row_number, salary_register, file_doc, error):
    log_entry = frappe.new_doc('DFM HR Log')
    log_entry.file_name = file_name
    log_entry.file_path = file_doc.file_url
    log_entry.row_number = row_number
    log_entry.status = "Partial Success"
    log_entry.dfm_hr_salary_transaction_summary = salary_register
    log_entry.date = frappe.utils.now_datetime()
    log_entry.error = error
    log_entry.insert(ignore_permissions=True)


def log_error(file_name, row_number, file_doc, error):
    log_entry = frappe.new_doc('DFM HR Log')
    log_entry.file_name = file_name
    log_entry.file_path = file_doc.file_url
    log_entry.row_number = row_number
    log_entry.status = "Error"
    log_entry.date = frappe.utils.now_datetime()
    log_entry.error = error
    log_entry.insert(ignore_permissions=True)







